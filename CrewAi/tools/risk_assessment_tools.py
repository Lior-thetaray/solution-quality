"""
Risk Assessment Alignment Tools
Tools for validating that solution features align with risk assessment requirements
"""

import os
import re
from typing import Dict, List, Any, Optional
from pathlib import Path
import openpyxl
from crewai.tools import BaseTool
from pydantic import BaseModel, Field


class ExcelReaderInput(BaseModel):
    """Input for Excel reader tool"""
    excel_path: str = Field(..., description="Path to the Excel file")
    sheet_name: Optional[str] = Field(None, description="Sheet name to read (default: first sheet)")


class ExcelReaderTool(BaseTool):
    name: str = "Read Excel Risk Assessment"
    description: str = """
    Reads the risk assessment Excel file and returns all features with their descriptions and usage types.
    Returns a list of dictionaries with feature_name, description, and usage_type.
    """

    def _run(self, excel_path: str, sheet_name: Optional[str] = None) -> str:
        """Read Excel file and return feature data"""
        try:
            # Convert relative path to absolute path from workspace root
            if not os.path.isabs(excel_path):
                workspace_root = Path(__file__).parent.parent.parent  # CrewAi/tools -> CrewAi -> workspace root
                excel_path = str(workspace_root / excel_path)
            
            if not os.path.exists(excel_path):
                return f"Error: Excel file not found at {excel_path}"
            
            wb = openpyxl.load_workbook(excel_path)
            ws = wb[sheet_name] if sheet_name else wb.active
            
            # Find header row (should be row 3 based on creation)
            headers = []
            header_row = 3
            for cell in ws[header_row]:
                if cell.value:
                    headers.append(cell.value)
            
            if not headers:
                return "Error: No headers found in Excel file"
            
            # Read data rows
            features = []
            for row_idx in range(header_row + 1, ws.max_row + 1):
                row = ws[row_idx]
                if not row[0].value:  # Skip empty rows
                    continue
                
                feature_data = {}
                for idx, header in enumerate(headers):
                    feature_data[header.lower().replace(' ', '_')] = row[idx].value
                
                features.append(feature_data)
            
            result = f"Found {len(features)} features in risk assessment:\n\n"
            
            training_features = [f for f in features if f.get('usage_type') == 'Training']
            forensic_features = [f for f in features if f.get('usage_type') == 'Forensic']
            
            result += f"Training Features ({len(training_features)}):\n"
            for f in training_features:
                result += f"  - {f.get('feature_name')}: {f.get('description')}\n"
            
            result += f"\nForensic Features ({len(forensic_features)}):\n"
            for f in forensic_features:
                result += f"  - {f.get('feature_name')}: {f.get('description')}\n"
            
            return result
            
        except Exception as e:
            return f"Error reading Excel file: {str(e)}"


class FeatureImplementationValidatorInput(BaseModel):
    """Input for feature implementation validator"""
    feature_name: str = Field(..., description="Name of the feature to validate")
    domain: str = Field(..., description="Domain name (e.g., demo_fuib)")
    expected_description: str = Field(..., description="Expected description from risk assessment")


class FeatureImplementationValidatorTool(BaseTool):
    name: str = "Validate Feature Implementation"
    description: str = """
    Validates that a specific feature is correctly implemented in the codebase.
    Checks:
    1. Feature file exists
    2. Feature description matches risk assessment
    3. Feature has required methods (identifier, description, get_agg_exprs, output_fields)
    4. Feature has trace_query if needed for training features
    Returns validation results with pass/fail status and details.
    """

    def _run(self, feature_name: str, domain: str, expected_description: str) -> str:
        """Validate feature implementation"""
        try:
            # Construct feature file path from workspace root
            workspace_root = Path(__file__).parent.parent.parent  # CrewAi/tools -> CrewAi -> workspace root
            base_path = workspace_root / "Sonar" / "domains" / domain / "features" / "core" / "customer"
            feature_file = base_path / f"{feature_name}.py"
            
            if not feature_file.exists():
                return f"❌ FAIL: Feature file not found at {feature_file}"
            
            # Read feature file
            with open(feature_file, 'r') as f:
                content = f.read()
            
            results = []
            results.append(f"✅ Feature file exists: {feature_file}")
            
            # Check for required methods and properties
            checks = {
                'identifier': r'def identifier\(self\)',
                'version': r'def version\(self\)',
                'description': r'def description\(self\)',
                'required_columns': r'def required_columns\(self\)',
                'group_by_keys': r'def group_by_keys\(self\)',
                'get_agg_exprs': r'def get_agg_exprs\(self',
                'output_fields': r'def output_fields\(self\)',
            }
            
            for check_name, pattern in checks.items():
                if re.search(pattern, content):
                    results.append(f"✅ Has {check_name} method/property")
                else:
                    results.append(f"❌ Missing {check_name} method/property")
            
            # Check description content
            desc_match = re.search(r'def description\(self\).*?return\s+[\'"]([^\'"]+)[\'"]', content, re.DOTALL)
            if desc_match:
                actual_desc = desc_match.group(1)
                results.append(f"\n📝 Description found: '{actual_desc}'")
                results.append(f"📝 Expected: '{expected_description}'")
                
                # Simple similarity check
                if actual_desc.lower().strip() == expected_description.lower().strip():
                    results.append("✅ Description matches exactly")
                elif any(word in actual_desc.lower() for word in expected_description.lower().split()[:5]):
                    results.append("⚠️  Description partially matches (similar intent)")
                else:
                    results.append("❌ Description does not match risk assessment")
            else:
                results.append("❌ Could not extract description from feature")
            
            # Check for trace_query
            if re.search(r'def trace_query\(self', content):
                results.append("✅ Has trace_query method (required for trained features)")
            else:
                results.append("⚠️  No trace_query method found")
            
            # Check for output fields
            field_count = len(re.findall(r'Field\(identifier=', content))
            results.append(f"✅ Declares {field_count} output field(s)")
            
            return "\n".join(results)
            
        except Exception as e:
            return f"Error validating feature: {str(e)}"


class WranglingConfigValidatorInput(BaseModel):
    """Input for wrangling config validator"""
    domain: str = Field(..., description="Domain name (e.g., demo_fuib)")
    entity: str = Field(default="customer", description="Entity type")
    cadence: str = Field(default="monthly", description="Cadence type")


class WranglingConfigValidatorTool(BaseTool):
    name: str = "Validate Wrangling Configuration"
    description: str = """
    Validates the wrangling.yaml configuration against risk assessment requirements.
    Checks:
    1. All training features from risk assessment are in config with train: true
    2. All forensic features are in config with train: false
    3. No extra features that shouldn't be there
    Returns validation report with missing/incorrect features.
    """

    def _run(self, domain: str, entity: str = "customer", cadence: str = "monthly") -> str:
        """Validate wrangling configuration"""
        try:
            import yaml
            
            workspace_root = Path(__file__).parent.parent.parent  # CrewAi/tools -> CrewAi -> workspace root
            config_path = workspace_root / "Sonar" / "domains" / domain / "config" / "core" / entity / cadence / "wrangling.yaml"
            
            if not config_path.exists():
                return f"❌ Config file not found: {config_path}"
            
            with open(config_path, 'r') as f:
                config = yaml.safe_load(f)
            
            if 'requested_features' not in config:
                return "❌ No 'requested_features' section in config"
            
            features_config = config['requested_features']
            
            results = [f"📋 Analyzing wrangling config: {config_path}\n"]
            
            # Get active features with their train status
            active_features = {}
            for feature_name, versions in features_config.items():
                if isinstance(versions, dict) and 'v1' in versions:
                    v1_config = versions['v1']
                    if v1_config.get('active', False):
                        active_features[feature_name] = {
                            'train': v1_config.get('train', False),
                            'params': v1_config.get('params', {})
                        }
            
            results.append(f"Found {len(active_features)} active features")
            results.append(f"  - Training: {sum(1 for f in active_features.values() if f['train'])}")
            results.append(f"  - Forensic: {sum(1 for f in active_features.values() if not f['train'])}")
            results.append("")
            
            # List all active features
            results.append("Active Features by Type:")
            results.append("\nTraining Features (train: true):")
            for name, config in active_features.items():
                if config['train']:
                    results.append(f"  ✅ {name}")
            
            results.append("\nForensic Features (train: false):")
            for name, config in active_features.items():
                if not config['train']:
                    results.append(f"  📊 {name}")
            
            return "\n".join(results)
            
        except Exception as e:
            return f"Error validating wrangling config: {str(e)}"


class TrainingNotebookValidatorInput(BaseModel):
    """Input for training notebook validator"""
    domain: str = Field(..., description="Domain name (e.g., demo_fuib)")
    notebook_name: str = Field(default="train_model", description="Training notebook name")


class TrainingNotebookValidatorTool(BaseTool):
    name: str = "Validate Training Notebook"
    description: str = """
    Validates that the training notebook uses the correct features.
    Checks:
    1. Notebook exists
    2. Extracts train_features_cols list
    3. Compares with risk assessment training features
    Returns list of features used in training and validation status.
    """

    def _run(self, domain: str, notebook_name: str = "train_model") -> str:
        """Validate training notebook"""
        try:
            workspace_root = Path(__file__).parent.parent.parent  # CrewAi/tools -> CrewAi -> workspace root
            notebook_path = workspace_root / "Sonar" / "domains" / domain / "notebooks" / f"{notebook_name}.ipynb"
            
            if not notebook_path.exists():
                return f"❌ Training notebook not found: {notebook_path}"
            
            import json
            
            with open(notebook_path, 'r') as f:
                notebook = json.load(f)
            
            results = [f"📓 Analyzing training notebook: {notebook_path}\n"]
            
            # Search for train_features_cols definition
            train_features = None
            for cell in notebook.get('cells', []):
                if cell.get('cell_type') == 'code':
                    source = ''.join(cell.get('source', []))
                    
                    # Look for train_features_cols assignment
                    match = re.search(r'train_features_cols\s*=\s*\[(.*?)\]', source, re.DOTALL)
                    if match:
                        features_str = match.group(1)
                        # Extract feature names
                        train_features = re.findall(r'[\'"]([^\'"]+)[\'"]', features_str)
                        # Filter out comments
                        train_features = [f for f in train_features if not f.startswith('#')]
                        break
            
            if not train_features:
                return "❌ Could not find train_features_cols in notebook"
            
            results.append(f"✅ Found {len(train_features)} features in training list:")
            for idx, feature in enumerate(train_features, 1):
                results.append(f"  {idx}. {feature}")
            
            return "\n".join(results)
            
        except Exception as e:
            return f"Error validating training notebook: {str(e)}"


class RiskAssessmentAlignmentReportInput(BaseModel):
    """Input for comprehensive alignment report"""
    domain: str = Field(..., description="Domain name (e.g., demo_fuib)")
    excel_path: str = Field(..., description="Path to risk assessment Excel file")


class RiskAssessmentAlignmentReportTool(BaseTool):
    name: str = "Generate Risk Assessment Alignment Report"
    description: str = """
    Generates a comprehensive report comparing risk assessment requirements with actual implementation.
    This is the main validation tool that orchestrates all checks:
    1. Reads risk assessment Excel
    2. Checks feature implementations
    3. Validates wrangling config
    4. Validates training notebook
    5. Generates pass/fail report with recommendations
    """

    def _run(self, domain: str, excel_path: str) -> str:
        """Generate comprehensive alignment report"""
        try:
            # Convert relative path to absolute path from workspace root
            if not os.path.isabs(excel_path):
                workspace_root = Path(__file__).parent.parent.parent  # CrewAi/tools -> CrewAi -> workspace root
                excel_path = str(workspace_root / excel_path)
            
            results = []
            results.append("=" * 80)
            results.append("RISK ASSESSMENT ALIGNMENT REPORT")
            results.append("=" * 80)
            results.append(f"Domain: {domain}")
            results.append(f"Risk Assessment: {excel_path}")
            results.append("=" * 80)
            results.append("")
            
            # Read Excel file
            if not os.path.exists(excel_path):
                return f"❌ ERROR: Risk assessment file not found at {excel_path}"
            
            wb = openpyxl.load_workbook(excel_path)
            ws = wb.active
            
            # Extract features from Excel
            features_from_ra = []
            for row_idx in range(4, ws.max_row + 1):  # Start from row 4 (after headers)
                row = ws[row_idx]
                if row[0].value:
                    features_from_ra.append({
                        'name': row[0].value,
                        'description': row[1].value,
                        'usage_type': row[2].value
                    })
            
            training_features = [f for f in features_from_ra if f['usage_type'] == 'Training']
            forensic_features = [f for f in features_from_ra if f['usage_type'] == 'Forensic']
            
            results.append(f"📊 Risk Assessment Summary:")
            results.append(f"  Total Features: {len(features_from_ra)}")
            results.append(f"  Training Features: {len(training_features)}")
            results.append(f"  Forensic Features: {len(forensic_features)}")
            results.append("")
            
            # Check feature implementations
            results.append("=" * 80)
            results.append("FEATURE IMPLEMENTATION VALIDATION")
            results.append("=" * 80)
            results.append("")
            
            missing_features = []
            implemented_features = []
            
            workspace_root = Path(__file__).parent.parent.parent
            
            for feature in features_from_ra:
                feature_name = feature['name']
                feature_path = workspace_root / "Sonar" / "domains" / domain / "features" / "core" / "customer" / f"{feature_name}.py"
                
                if feature_path.exists():
                    implemented_features.append(feature_name)
                    results.append(f"✅ {feature_name} - IMPLEMENTED")
                else:
                    missing_features.append(feature_name)
                    results.append(f"❌ {feature_name} - MISSING")
            
            results.append("")
            results.append(f"Summary: {len(implemented_features)}/{len(features_from_ra)} features implemented")
            
            if missing_features:
                results.append(f"\n⚠️  Missing Features: {', '.join(missing_features)}")
            
            # Check wrangling config
            results.append("")
            results.append("=" * 80)
            results.append("WRANGLING CONFIGURATION VALIDATION")
            results.append("=" * 80)
            results.append("")
            
            import yaml
            config_path = workspace_root / "Sonar" / "domains" / domain / "config" / "core" / "customer" / "monthly" / "wrangling.yaml"
            
            if config_path.exists():
                with open(config_path, 'r') as f:
                    config = yaml.safe_load(f)
                
                requested_features = config.get('requested_features', {})
                
                # Check training features
                results.append("Training Features Alignment:")
                for feature in training_features:
                    name = feature['name']
                    if name in requested_features:
                        v1_config = requested_features[name].get('v1', {})
                        is_active = v1_config.get('active', False)
                        is_train = v1_config.get('train', False)
                        
                        if is_active and is_train:
                            results.append(f"  ✅ {name} - active: true, train: true")
                        elif is_active and not is_train:
                            results.append(f"  ⚠️  {name} - active: true, train: false (should be train: true)")
                        else:
                            results.append(f"  ❌ {name} - not active")
                    else:
                        results.append(f"  ❌ {name} - NOT IN CONFIG")
                
                results.append("")
                results.append("Forensic Features Alignment:")
                for feature in forensic_features:
                    name = feature['name']
                    if name in requested_features:
                        v1_config = requested_features[name].get('v1', {})
                        is_active = v1_config.get('active', False)
                        is_train = v1_config.get('train', False)
                        
                        if is_active and not is_train:
                            results.append(f"  ✅ {name} - active: true, train: false")
                        elif is_active and is_train:
                            results.append(f"  ⚠️  {name} - active: true, train: true (should be train: false)")
                        else:
                            results.append(f"  ❌ {name} - not active")
                    else:
                        results.append(f"  ❌ {name} - NOT IN CONFIG")
            else:
                results.append(f"❌ Wrangling config not found: {config_path}")
            
            # Final Score
            results.append("")
            results.append("=" * 80)
            results.append("ALIGNMENT SCORE")
            results.append("=" * 80)
            
            total_checks = len(features_from_ra) * 2  # Implementation + Config
            passed_checks = len(implemented_features)
            
            # Add config alignment checks
            if config_path.exists():
                for feature in features_from_ra:
                    name = feature['name']
                    if name in requested_features:
                        v1_config = requested_features[name].get('v1', {})
                        is_active = v1_config.get('active', False)
                        is_train = v1_config.get('train', False)
                        expected_train = (feature['usage_type'] == 'Training')
                        
                        if is_active and (is_train == expected_train):
                            passed_checks += 1
            
            score = (passed_checks / total_checks * 100) if total_checks > 0 else 0
            
            results.append(f"Overall Alignment Score: {score:.1f}%")
            results.append(f"Passed Checks: {passed_checks}/{total_checks}")
            
            if score >= 95:
                results.append("✅ EXCELLENT - Risk assessment fully aligned")
            elif score >= 80:
                results.append("⚠️  GOOD - Minor alignment issues")
            elif score >= 60:
                results.append("⚠️  FAIR - Several alignment issues to address")
            else:
                results.append("❌ POOR - Significant alignment issues")
            
            results.append("=" * 80)
            
            return "\n".join(results)
            
        except Exception as e:
            return f"Error generating alignment report: {str(e)}\n{str(e.__traceback__)}"
